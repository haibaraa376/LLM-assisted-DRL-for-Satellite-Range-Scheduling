"""从Excel清单安全登记、匹配、整理和回滚PDF来源。"""

import csv
from datetime import datetime, timezone
import hashlib
import json
from pathlib import Path
import shutil
import uuid

from openpyxl import load_workbook

from .title_matching import match_document


EXCLUDED_DIRECTORIES = {
    "original_titles",
    "raw",
    "processed",
    "index",
    "reports",
    "review",
}
CATEGORY_DIRECTORIES = {
    "卫星测距调度": "satellite_scheduling",
    "卫星任务调度": "satellite_scheduling",
    "卫星网络资源调度": "satellite_scheduling",
    "卫星地面站调度": "satellite_scheduling",
    "奖励塑形": "reward_shaping",
    "多智能体强化学习": "reward_shaping",
    "LLM奖励设计": "reward_shaping",
    "RAG实现": "implementation_reference",
    "向量化实现": "implementation_reference",
}
REQUIRED_COLUMNS = (
    "文献ID", "类别", "优先级", "纳入知识库", "题目", "作者", "年份",
    "期刊/会议", "DOI", "来源URL", "为什么需要", "建议抽取内容", "获取情况",
    "当前状态", "建议文件名", "备注",
)


def sha256_file(path):
    """流式计算文件SHA256，不将PDF整体载入内存。"""
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def atomic_write_json(path, payload):
    """原子写入JSON，防止报告中途截断。"""
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_suffix(target.suffix + ".tmp")
    temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(target)


def load_spreadsheet_records(path):
    """只以“文献清单”工作表作为命名和元数据来源。"""
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    if "文献清单" not in workbook.sheetnames:
        raise ValueError("Excel缺少文献清单工作表")
    rows = workbook["文献清单"].iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    missing = set(REQUIRED_COLUMNS) - set(headers)
    if missing:
        raise ValueError("Excel缺少必要列：{0}".format(sorted(missing)))
    records = []
    for values in rows:
        record = {
            header: "" if value is None else str(value).strip()
            for header, value in zip(headers, values)
        }
        if record["文献ID"]:
            records.append(record)
    if len({item["文献ID"] for item in records}) != len(records):
        raise ValueError("Excel存在重复文献ID")
    return records


def scan_pdf_files(root):
    """递归扫描未整理区域的PDF，避免生成目录二次入库。"""
    root = Path(root)
    return sorted(
        path for path in root.rglob("*.pdf")
        if not any(part in EXCLUDED_DIRECTORIES for part in path.relative_to(root).parts)
    )


def _metadata_title(path):
    """尽力读取PDF元数据标题；失败不影响安全匹配。"""
    try:
        import fitz
        document = fitz.open(path)
        title = (document.metadata or {}).get("title", "")
        document.close()
        return title or ""
    except Exception:
        return ""


def build_match_report(root, spreadsheet, threshold=0.92, margin=0.08):
    """构建不修改PDF的匹配报告与选择核对结果。"""
    records = load_spreadsheet_records(spreadsheet)
    root = Path(root)
    matches = []
    claimed = {}
    for path in scan_pdf_files(root):
        record, method, score, second, reason = match_document(
            path,
            records,
            [path.stem, _metadata_title(path)],
            threshold,
            margin,
        )
        document_id = record["文献ID"] if record else ""
        status = "matched" if record else "needs_manual_resolution"
        if document_id in claimed:
            status = "needs_manual_resolution"
            reason = "同一文献ID被多个PDF匹配"
        if document_id:
            claimed.setdefault(document_id, []).append(str(path))
        matches.append({
            "original_path": str(path.relative_to(root)),
            "original_filename": path.name,
            "matched_document_id": document_id,
            "matched_title": record["题目"] if record else "",
            "suggested_filename": record["建议文件名"] if record else "",
            "category": record["类别"] if record else "",
            "include_in_kb": record["纳入知识库"] if record else "",
            "match_method": method,
            "match_score": round(score, 6),
            "second_best_score": round(second, 6),
            "status": status,
            "reason": reason,
            "sha256": sha256_file(path),
        })
    duplicates = [
        {"document_id": key, "paths": value}
        for key, value in claimed.items() if len(value) > 1
    ]
    matched_ids = {item["matched_document_id"] for item in matches if item["status"] == "matched"}
    expected = [item for item in records if item["纳入知识库"] == "是"]
    reconciliation = {
        "pdf_file_count": len(matches),
        "spreadsheet_row_count": len(records),
        "spreadsheet_a_core_count": sum(item["优先级"] == "A级核心" for item in records),
        "spreadsheet_include_yes_count": len(expected),
        "matched_file_count": len(matched_ids),
        "unmatched_files": [item["original_filename"] for item in matches if item["status"] != "matched"],
        "missing_expected_documents": [item["文献ID"] for item in expected if item["文献ID"] not in matched_ids],
        "duplicate_matches": duplicates,
        "notes": ["Excel是唯一元数据来源；未下载条目不会创建伪文件。"],
    }
    return records, matches, reconciliation


def write_match_reports(root, matches, reconciliation):
    """写入JSON、CSV和选择核对报告。"""
    reports = Path(root) / "reports"
    reports.mkdir(parents=True, exist_ok=True)
    atomic_write_json(reports / "title_match_report.json", matches)
    atomic_write_json(reports / "selection_reconciliation.json", reconciliation)
    with (reports / "title_match_report.csv").open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(matches[0]) if matches else ["original_filename"])
        writer.writeheader()
        writer.writerows(matches)


def apply_operations(root, matches):
    """保留原始标题PDF，并复制规范副本；任何冲突均不覆盖。"""
    root = Path(root)
    manifest = root / "reports" / "file_operations.jsonl"
    operations = []
    for item in matches:
        if item["status"] != "matched":
            continue
        category = CATEGORY_DIRECTORIES.get(item["category"])
        if category is None:
            raise ValueError("未知类别，拒绝自动整理：{0}".format(item["category"]))
        source = root / item["original_path"]
        original_destination = root / "original_titles" / source.name
        canonical_destination = root / "raw" / category / item["suggested_filename"]
        if canonical_destination.exists():
            same = sha256_file(canonical_destination) == item["sha256"]
            operations.append({"operation_id": str(uuid.uuid4()), "timestamp": datetime.now(timezone.utc).isoformat(), "source_path": str(source), "destination_path": str(canonical_destination), "operation": "copy2", "source_sha256": item["sha256"], "destination_sha256": sha256_file(canonical_destination), "status": "already_present_same_content" if same else "conflict_different_content"})
            if not same:
                continue
        else:
            original_destination.parent.mkdir(parents=True, exist_ok=True)
            if not original_destination.exists():
                shutil.move(str(source), str(original_destination))
            elif sha256_file(original_destination) != item["sha256"]:
                raise FileExistsError("original_titles存在同名不同内容文件")
            canonical_destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(original_destination, canonical_destination)
            operations.append({"operation_id": str(uuid.uuid4()), "timestamp": datetime.now(timezone.utc).isoformat(), "source_path": str(source), "destination_path": str(canonical_destination), "operation": "move_then_copy2", "source_sha256": item["sha256"], "destination_sha256": sha256_file(canonical_destination), "status": "created"})
    with manifest.open("a", encoding="utf-8") as stream:
        for operation in operations:
            stream.write(json.dumps(operation, ensure_ascii=False) + "\n")
    return operations


def write_source_register(root, records, matches):
    """根据已匹配PDF和Excel唯一元数据建立初始来源登记表。"""
    root = Path(root)
    by_id = {record["文献ID"]: record for record in records}
    output = []
    for match in matches:
        if match["status"] != "matched":
            continue
        record = by_id[match["matched_document_id"]]
        directory = CATEGORY_DIRECTORIES[record["类别"]]
        output.append({
            "document_id": record["文献ID"], "title": record["题目"],
            "authors": record["作者"], "year": record["年份"],
            "venue": record["期刊/会议"], "doi": record["DOI"],
            "source_url": record["来源URL"], "category": directory,
            "priority": record["优先级"],
            "include_in_knowledge_base": record["纳入知识库"] == "是",
            "researcher_approved": True,
            "technical_extraction_status": "pending_technical_check",
            "original_filename": match["original_filename"],
            "canonical_filename": record["建议文件名"],
            "canonical_relative_path": str(Path("raw") / directory / record["建议文件名"]),
            "sha256": match["sha256"],
            "file_size_bytes": (root / "raw" / directory / record["建议文件名"]).stat().st_size,
            "page_count": None, "why_needed": record["为什么需要"],
            "recommended_content": record["建议抽取内容"], "notes": record["备注"],
        })
    csv_path = root / "source_register.csv"
    jsonl_path = root / "source_register.jsonl"
    fields = list(output[0]) if output else ["document_id"]
    with csv_path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader(); writer.writerows(output)
    jsonl_path.write_text("\n".join(json.dumps(item, ensure_ascii=False) for item in output) + "\n", encoding="utf-8")
    return output
